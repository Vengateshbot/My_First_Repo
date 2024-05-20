# Author : VENGATESH D
# This script is a draft version.
# This will automate the build release checkout using the info from build logs

import shutil
import subprocess
from subprocess import *
import multiprocessing
import threading
import time, os, msvcrt
from win32com.client import *
from win32com.client.connect import *
import logging
import pythoncom
import win32com
import sys
import platform
import pathlib
from pathlib import Path
import ctypes        # module for C data types
import enum
import psutil
import zipfile
from py_canoe import CANoe 
import openpyxl
from datetime import datetime
from timeit import default_timer as timer
from cryptography.hazmat.primitives.ciphers import Cipher, algorithms, modes
from cryptography.hazmat.backends import default_backend
from cryptography.hazmat.primitives import padding
from cryptography.hazmat.primitives import hashes


def zip_folders(root_dir):
  """Zips each folder in the root directory into a separate zip file."""

  for root, dirs, files in os.walk(root_dir):
    for dir_name in dirs:
      # Create zip file with path relative to root directory
      zip_file_name = os.path.join(root_dir, dir_name + ".zip")

      with zipfile.ZipFile(zip_file_name, "w", zipfile.ZIP_DEFLATED) as zipf:
        for root_path, folders, files in os.walk(os.path.join(root, dir_name)):
          for file in files:
            relative_path = os.path.relpath(os.path.join(root_path, file), root_dir)
            zipf.write(os.path.join(root_path, file), relative_path)

      print(f"Created zip file: {zip_file_name}")


def copy_folders_and_files(source_path, destination_path):

  for root, dirs, files in os.walk(source_path):
    relative_path = os.path.relpath(root, source_path)  
    destination_dir = os.path.join(destination_path, relative_path)

    
    os.makedirs(destination_dir, exist_ok=True)

    for file in files:
      source_file = os.path.join(root, file)
      destination_file = os.path.join(destination_dir, file)
      shutil.copy2(source_file, destination_file) 

def log_command_output(file_path,arguments, log_file):

  os.chdir(file_path)  
  batch_file_name = "build.bat"	
  command = f"{batch_file_name} {' '.join(arguments)}"	
  try:
    with open(log_file, 'w') as file:
       
      completed_process = subprocess.run(command, stdout=subprocess.PIPE, stderr=subprocess.STDOUT,text=True,shell=True,check=True)
      # Write captured output to the file
      print(completed_process.stdout)
      file.write(completed_process.stdout)
      print("Batch file executed successfully.\n")
      print(f"Command output logged to: {log_file}")
      file.close()
      if completed_process.returncode != 0:
        print(f"Warning: Command exited with code {completed_process.returncode}")
  except subprocess.CalledProcessError as error:
    print(f"Error executing command: {error}")


def execute_batch_file(file_path, batch_file_name, arguments):

    os.chdir(file_path)  

    command = f"{batch_file_name} {' '.join(arguments)}" 
	
    try:
        subprocess.run(command, check=True, shell=True) 
        print("Batch file executed successfully.\n")
    except subprocess.CalledProcessError as error:
        print("Error executing batch file:\n", error)
		
		
def find_string_in_file(filename, string_to_find):

    if os.path.exists(filename):
        with open(filename, 'r') as file:
            contents = file.read()
            return string_to_find in contents
    else:
        
        return False

def delete_folder_contents(folder_path):

    for item in os.listdir(folder_path):
        item_path = os.path.join(folder_path, item)
        if os.path.isfile(item_path):
            os.remove(item_path)
        elif os.path.isdir(item_path):
            shutil.rmtree(item_path)

def delete_folder(folder_path):

    try:
        if os.path.exists(folder_path):
            if os.path.isdir(folder_path):
                shutil.rmtree(folder_path)  
            else:
                os.remove(folder_path)  # Delete a single file
            print(f"Folder '{folder_path}' deleted successfully!\n")
        else:
            print(f"Folder '{folder_path}' not found.")
    except OSError as e:
        print(f"Error deleting folder: {e}")

def create_folder(path, folder_name):

    try:
        os.makedirs(os.path.join(path, folder_name))
        print(f"Folder '{folder_name}' created successfully at '{path}'!\n")
    except OSError as error:
        print(f"Error creating folder: {error}")


def copy_files(source_folders, destination_folder, filename_list):


    for source_folder in source_folders:
        for root, _, files in os.walk(source_folder):
            for filename in files:
                if filename in filename_list:
                    source_file = os.path.join(root, filename)
                    destination_file = os.path.join(destination_folder, filename)
                    shutil.copy2(source_file, destination_file)
                    print(f"Copied {filename} to {destination_folder}")
                    
                    
def write_to_excel(row_i,coloumn_i,data_to_write,sheet_no):

	Work_book_path = r"C:\Flashing\TEST\Test_Excel.xlsx"
	workbook = openpyxl.load_workbook(Work_book_path)
	global worksheet1
	worksheet1 = workbook.worksheets[sheet_no] #workbook["Sheet1"] 
	
	worksheet1.cell(row=row_i, column=coloumn_i).value = data_to_write
	
	
	workbook.save(Work_book_path)
	
	workbook.close()
	
	
def read_from_excel(row_i,coloumn_i,Sheet_no):

	Work_book_path = r"C:\Flashing\TEST\Test_Excel.xlsx"
	workbook = openpyxl.load_workbook(Work_book_path)
	global worksheet2
	worksheet2 = workbook.worksheets[Sheet_no] #workbook["Sheet2"] 
	
	data_from_excel = worksheet2.cell(row=row_i, column=coloumn_i).value 
	
	
	workbook.save(Work_book_path)
	
	workbook.close()
	return data_from_excel


def Clear_excel_Report(Sheet_no):

	Work_book_path = r"C:\Flashing\TEST\Test_Excel.xlsx"
	workbook = openpyxl.load_workbook(Work_book_path)
	global worksheet1
	worksheet1 = workbook.worksheets[Sheet_no] #workbook["Sheet1"] 
	
	for col in worksheet1.iter_cols(min_col=2, max_col=2):
		for cell in col:
			cell.value = None
	
	
	workbook.save(Work_book_path)
	workbook.close()
    
def Ram_Rom_usage():
	global RAM_usage,FLASH_usage
	filename = r"C:\VENGATESH\10034766_SW_Honda_SRR6P\out\srr6p\srr6p_memory_stats.txt"  
	line1 = 19   # 19 for line 20
	line2 = 31  # 31 for line 32
	try:
		with open(filename, 'r') as file:
			lines = file.readlines()
			RAM_usage = lines[line1].strip("Total:  ")
			FLASH_usage = lines[line2].strip("Total:  ")
			write_to_excel(9,2,RAM_usage,0)
			write_to_excel(10,2,FLASH_usage,0)
		   
	except FileNotFoundError:
		print(f"Error: The file '{filename}' was not found.")
	except IndexError:
		print(f"Error: The requested line numbers are out of range for the file '{filename}'.")
        


def replace_lines(file_path, replacements):
	try:
	
		with open(file_path, 'r') as file:
			lines = file.readlines()

		for line_number, replacement_string in replacements.items():

			if (1 <= line_number <= len(lines)):
			
				lines[line_number - 1] = replacement_string + '\n'

				print(f"Line {line_number} replaced successfully.")
			else:
				print(f"Line {line_number} not replaced successfully.")
				
		with open(file_path, 'w') as file:
			file.writelines(lines)

	except FileNotFoundError:
		print(f"File not found: {file_path}")
	except Exception as e:
		print(f"An error occurred: {e}")

	

def DebuggerGetVal(variable_name,DebuggerVarName,capture_delay,loglist_enable):

	variable_name = ctypes.c_int32(0)
	formatted_string = f"VAR.ADDWATCH.{DebuggerVarName}"
	byte_string = formatted_string.encode("utf-8")
	t32api.T32_Cmd(byte_string)
	time.sleep(capture_delay)

	formatted_string = f"EVAL Var.VALUE({DebuggerVarName})"
	byte_string = formatted_string.encode("utf-8")
	t32api.T32_Cmd(byte_string)
	t32api.T32_EvalGet(ctypes.byref(variable_name))
	if(loglist_enable == 1):
		loglist.append(variable_name.value)	
	time.sleep(2)
	return variable_name.value
	


def Comment_Make_File(makefile_path, search_string, is_full_build):

    try:
        # Read the content of the make file
        with open(makefile_path, 'r') as file:
            lines = file.readlines()

        # Find the index of the line containing the search string
        index_of_search_string = next((index for index, line in enumerate(lines) if search_string in line), None)

        if index_of_search_string is not None:
            # Check if the lines are already commented or not
            lines_to_comment = lines[index_of_search_string + 1:min(index_of_search_string + 8, len(lines))]# Check if the lines are already commented or not
            are_lines_commented = all(line.startswith('#') for line in lines_to_comment)

            
            for i in range(index_of_search_string + 1, min(index_of_search_string + 8, len(lines))):
                if is_full_build and are_lines_commented:
                    lines[i] = lines[i][1:]  # Uncomment the line
                elif not is_full_build and not are_lines_commented:
                    lines[i] = '#' + lines[i]  # Comment the line

            
            with open(makefile_path, 'w') as file:
                file.writelines(lines)

            print("Successfully toggled commenting for the next 7 lines after '{}' in the make file.".format(search_string))
        else:
            print("The specified string '{}' was not found in the make file.".format(search_string))

    except FileNotFoundError:
        print("Error: Make file '{}' not found.".format(makefile_path))
    except Exception as e:
        print("An error occurred: {}".format(str(e)))

	
def close_applications(app_list):
	for app_name in app_list:
		for proc in psutil.process_iter(['pid', 'name']):
			if app_name.lower() in proc.info['name'].lower():
					try:
						pid = proc.info['pid']
						os.system(f"taskkill /F /PID {pid}")
						print(f"Closed {app_name} (PID: {pid})")
					except psutil.NoSuchProcess:
						print(f"Process {app_name} not found.")
					except psutil.AccessDenied:
						print(f"Access denied to terminate {app_name}.")
					except Exception as e:
						print(f"Error terminating {app_name}: {e}")




def Rename_Excel():

	current_file_path = f"C:/VENGATESH/Release_Folder/{Final_SW_folder_name}/RCR_HONDA.xlsx" 
	new_filename = f"RCR_HONDA_{SW_Version_canoe}.xlsx"

	new_file_path = os.path.join(os.path.dirname(current_file_path), new_filename)

	try:
	  
	  os.rename(current_file_path, new_file_path)
	  print("File renamed successfully!")
	except FileNotFoundError:
	  print("Error: File not found!")
	except PermissionError:
	  print("Error: Insufficient permissions to rename the file.")
	except Exception as e:
	  print(f"Unexpected error: {e}")


def Copy_Excel():
		
	source_path =  r"C:\Flashing\TEST\RCR_HONDA.xlsx"
	destination_path = f"C:/VENGATESH/Release_Folder/{Final_SW_folder_name}" 

	try:
	  # Copy the file
	  shutil.copy2(source_path, destination_path)
	  print(f"File copied successfully from {source_path} to {destination_path}")
	except FileNotFoundError:
	  print(f"Error: Source file '{source_path}' not found.")
	except PermissionError:
	  print(f"Error: Permission denied while copying the file.")
	except Exception as e:
	  print(f"An unexpected error occurred: {e}")


def Fill_Checklist(SW_version):
	now = datetime.now()
	Release_date = now.strftime("%d-%m-%Y")
	Release_type = "Engineering Drop"
	Work_book_path=r"C:\Flashing\TEST\RCR_HONDA.xlsx"
	workbook = openpyxl.load_workbook(Work_book_path)
	
	global Header
	Header = workbook.worksheets[0] #workbook["Sheet1"] 

	(Header.cell(row=7, column=2).value) = SW_version
	# cell_value = worksheet1["A27"].value
	i=0
	Row=14
	while(1):
		i=Header.cell(row=Row, column=1).value
		if(i==None):
			(Header.cell(row=Row, column=1).value) = int((Header.cell(row=Row-1, column=1).value)) + 1
			(Header.cell(row=Row, column=2).value) = Release_date
			(Header.cell(row=Row, column=3).value) = (Header.cell(row=Row-1, column=3).value)# build engineer name
			(Header.cell(row=Row, column=4).value) = Release_type
			(Header.cell(row=Row, column=5).value) = SW_version
			(Header.cell(row=Row, column=5).hyperlink) = "https://confluence.asux.aptiv.com/pages/resumedraft.action?draftId=249443163&draftShareId=6bd3f786-8b9d-4402-8933-b0d391844e14&"
			print("Checklist Filled")
			break
		Row=Row+1
	# # Access the cell value at row 2, column A
	# cell_value = worksheet1.cell(row=2, column=1).value # another way is worksheet1["A2"].value

	Release_information = workbook.worksheets[1]
	(Release_information.cell(row=6, column=2).value) = Release_date
	(Release_information.cell(row=12, column=2).value) = SW_version
	(Release_information.cell(row=13, column=2).value) = Release_date
	(Release_information.cell(row=5, column=2).value) = (Header.cell(row=Row, column=3).value)  # build engineer name
	
	Main_Checklist = workbook.worksheets[2]
	(Main_Checklist.cell(row=33, column=5).value) = f"RAM Utilization : {RAM_usage} \nROM Utilization : {FLASH_usage}"

	workbook.save(Work_book_path)
	workbook.close()



def Start_flashing(flash_delay):
    
	start = timer()
	global flash_state
	flash_state=0

	if platform.system() == "Windows":
		SYSDIR='C:/T32'
		if platform.machine().endswith('64'):
			#win64
			APIFILE='t32api64.dll'
		else:
			#win32
			APIFILE='t32api.dll'
	else:
		print("Unknown OS")
		quit()
		
	global t32api
	global LIBFILE
	APIDIR='C:/Flashing/TEST'
	LIBFILE=os.path.join(os.sep,SYSDIR,APIDIR,APIFILE)
	t32api=ctypes.cdll.LoadLibrary(LIBFILE)
	T32_DEV = 1
	t32api.T32_Config(b"NODE=",b"localhost")
	t32api.T32_Config(b"PORT=", b"20001")
	t32api.T32_Config(b"PACKLEN=",b"1024")

	directory='C:/Flashing/TEST/'
	file_name='Logs1.txt'
	Log_file_path = os.path.join(directory, file_name)


	batch_file_path = r'C:\Flashing\TEST\KillProcesses.bat'
    
    
    
    
	def Check_Disk_Space_available():
		partitions = psutil.disk_partitions()
		for partition in partitions:
			if partition.device.startswith('C:'):
				disk_usage = psutil.disk_usage(partition.mountpoint)
				disk_space_available=int(disk_usage.free/(1024**3))
				
		return disk_space_available  
        
				
                

    # -----------------------------------------------------------------------------
    # main
    # -----------------------------------------------------------------------------
	if (Check_Disk_Space_available() <=5): 
		sys.exit("Low Disk Space available") 
		
	rc1 = t32api.T32_Init()
	for y in range(0,3):
		rc1 = t32api.T32_Attach(T32_DEV)
		if rc1 == 0:
			break
	if rc1 != 0:
		t32api.T32_Exit()
		assert rc1 != 0 ,"T32 instance opened already" 
	rc1 = t32api.T32_Ping()
	if rc1 != 0:
		t32api.T32_Exit()
		assert rc1 != 0 ,"T32 instance opened already"

	time.sleep(2)
	t32api.T32_Cmd(b"QUIT")
 
    # loads the sample configuration
    # add test modules to the configuration
    #Start Lauterbach by executing corresponding bactch file which will execute corresponding .cmm file
	subprocess.run(r'C:\Flashing\TEST\_start_powerview_SMP.bat',check=True, shell=True,cwd=r'C:\Flashing\TEST')
	time.sleep(flash_delay)#90


	rc = t32api.T32_Init()
	for x in range(0,3):
		rc = t32api.T32_Attach(T32_DEV)
		if rc == 0:
			break
	if rc != 0:
	   t32api.T32_Exit()
	   assert (rc != 0),"Connection to Trace32 Failed"
	   sys.exit(1) 
	rc = t32api.T32_Ping()
	if rc != 0:
	   t32api.T32_Exit()
	   assert (rc != 0),"Connection to Trace32 Failed after Ping"
	   sys.exit(1) 
	
	global loglist,reset_count, countvalue_5ms
	loglist=[]
	t32api.T32_Cmd(b"Break.Delete /All")
	t32api.T32_Cmd(b"SYStem.Option DUALPORT ON")
	time.sleep(2)
	t32api.T32_Cmd(b"DO Quit SMP")
	time.sleep(25)
	t32api.T32_Cmd(b"GO")

	Count_5ms = ctypes.c_int32(0)
	t32api.T32_Cmd(b"VAR.ADDWATCH.cnt_5ms")
	time.sleep(30)
	t32api.T32_Cmd(b"EVAL Var.VALUE(cnt_5ms)")
	t32api.T32_EvalGet(ctypes.byref(Count_5ms))
	Valx=Count_5ms.value
	#print("5 ms count value = " + str(Count_5ms.value) + "\n")
	reset_count=0
	countvalue_5ms = 5000
	while(Valx < countvalue_5ms):
	
		t32api.T32_Cmd(b"system.resettarget")
		time.sleep(4)
		t32api.T32_Cmd(b"GO")
		reset_count = reset_count + 1
		t32api.T32_Cmd(b"Var.DelWatch.cnt_5ms")
		Valx=DebuggerGetVal("Count_5ms","cnt_5ms",30,0)
		if(Valx > countvalue_5ms):
			break
	loglist.append(Valx)
	write_to_excel(13,2,Valx,0)		
	write_to_excel(15,2,reset_count,0)
	
	Wakeup_status = ctypes.c_int8(0)
	t32api.T32_Cmd(b"VAR.ADDWATCH.Radar_Data.Init_Status")
	time.sleep(5)
	t32api.T32_Cmd(b"EVAL Var.VALUE(Radar_Data.Init_Status)")
	t32api.T32_EvalGet(ctypes.byref(Wakeup_status))
	#print("Wake up value = " + str(Wakeup_status.value) + "\n")
	loglist.append(Wakeup_status.value)
	write_to_excel(14,2,Wakeup_status.value,0)
	time.sleep(2)
	
	Can_busoff_fault = ctypes.c_int8(0)
	t32api.T32_Cmd(b"VAR.ADDWATCH.Customer_Active_Fault_Table.Customer_Bits.CAN_Busoff_Fault")
	time.sleep(2)
	t32api.T32_Cmd(b"EVAL Var.VALUE(Customer_Active_Fault_Table.Customer_Bits.CAN_Busoff_Fault)")
	t32api.T32_EvalGet(ctypes.byref(Can_busoff_fault))
	loglist.append(Can_busoff_fault.value)
	write_to_excel(4,2,Can_busoff_fault.value,0)
	time.sleep(2)

	if(Can_busoff_fault.value == 1):
		print("Can bus off fault is set\n")
		sys.exit(1)
        
	val1=DebuggerGetVal("release_version","Release_Revision",2,1)

	val2=DebuggerGetVal("Promote_version","Promote_Revision",2,1)

	val3=DebuggerGetVal("Field_version","SW_Field",2,1)

	sw_version = f"{val1}.{format(val2,'x').upper()}.{format(val3,'x').upper()}"
	print(sw_version)
	write_to_excel(2,2,sw_version,0)
    
	t32api.T32_Cmd(b"SYStem.mode down")
	# t32api.T32_Cmd(b"SYStem.mode up")
	# time.sleep(2)
	# t32api.T32_Cmd(b"QUIT")
	
	if((loglist[0] >= countvalue_5ms) and (loglist[1] == 3)):
			if os.path.exists(Log_file_path):
				with open(Log_file_path, 'r') as file:
					first_line = file.readline().strip()
					file.close()
					if(first_line == "ECU Wakeup Failed."):
						file1open=open(Log_file_path,'w')
						file1open.write("ECU Wakeup Successful.")
						file1open.close()
					elif(first_line == "ECU Wakeup Successful."):
						file2open=open(Log_file_path,'w')
						file2open.write("ECU Wakeup Successful.")
						file2open.close()
						

			else:
				assert "Log file not found"
	   
	elif((loglist[0] < countvalue_5ms) and (loglist[1] != 3)):
			if os.path.exists(Log_file_path):
				with open(Log_file_path, 'r') as file:
					first_line = file.readline().strip()
					file.close()
					if(first_line == "ECU Wakeup Failed."):
						file1open=open(Log_file_path,'w')
						file1open.write("ECU Wakeup Failed.")
						file1open.close()
					elif(first_line == "ECU Wakeup Successful."):
						file2open=open(Log_file_path,'w')
						file2open.write("ECU Wakeup Failed.")
						file2open.close()
						
			else:
				assert "Log file not found"
											
	content2 = open("C:\\Flashing\\TEST\\Logs.txt","r")
	content1 = open("C:\\Flashing\\TEST\\Logs1.txt","r")
	check1=[]
	check2=[]
	for line in content1:
		check1.append(line.strip()) 
	for line1 in content2:
		check2.append(line1.strip())
		
	content1.close()
	content2.close()  
	time.sleep(2)  
	if((((loglist[1] == 3) and (loglist[0] >=countvalue_5ms)) or (check1[0] == "ECU Wakeup Successful.")) and (check2[0] == "Flashing Successful.")):
		print("Success")
		end=timer()
		print("Execution time = " + str(end-start) + " Sec")
		flash_state=1
		
		
	elif(loglist[0] < countvalue_5ms):
		print("Fail-1")
		end=timer()
		print("Execution time = " + str(end-start) + " Sec")
		flash_state=0
		sys.exit(f"Counter 5 ms not updated to {countvalue_5ms}") 
		 
	else:
		print("Fail-2")
		end=timer()
		print("Execution time = " + str(end-start) + " Sec")
		flash_state=0
		sys.exit("Flashing is Success but ECU wakeup failed") 
		
def Start_Canoe():

	global canoe_state
	canoe_state=0
	def remove_bytes(hex_input):
		# Convert hexadecimal input to bytes
		byte_input = bytes.fromhex(hex_input)

		# Remove the first 2 bytes and last 2 bytes
		modified_bytes = byte_input[2:-2]

		# Convert the modified bytes back to hexadecimal
		modified_hex = modified_bytes.hex()

		return modified_hex
		
			
			
	def Type_V_Security(response_key):

		def circular_right_shift(num, shift):

			return (num >> shift) | (num << (16 - shift))

		def circular_left_shift(num, shift):

			return (num << shift) | (num >> (16 - shift))

			
		user_input_str = response_key

		# Convert to integer
		user_input = int(user_input_str, 16)
		value1 = circular_right_shift(user_input + 0xA6A, 8)
		value2 = circular_left_shift(user_input + 0x1C3D, 1)
		value3 = value1 ^ value2
		value4 = value3 + 0x181
		value4_bytes = value4.to_bytes(3, byteorder='big')  # Convert to 2-byte bytes object
		second_third_bytes = value4_bytes[1:]  # Extract 2nd and 3rd bytes
		final_key = f"2708{second_third_bytes.hex()}A184"  # Use hex() to format as hex string
		#print(f"Final_Key: {final_key}")
		
		return final_key
            
            
            
            
	def Type_X_Security(Response): 

		def remove_bytes(hex_input):
			# Convert hexadecimal input to bytes
			byte_input = bytes.fromhex(hex_input)

			# Remove the first 2 bytes and last 2 bytes
			modified_bytes = byte_input[2:-2]

			# Convert the modified bytes back to hexadecimal
			modified_hex = modified_bytes.hex()

			return modified_hex

		def pad_data(data):
			padder = padding.PKCS7(128).padder()
			padded_data = padder.update(data) + padder.finalize()
			return padded_data

		def encrypt(key, data):
			backend = default_backend()
			cipher = Cipher(algorithms.AES(key), modes.ECB(), backend=backend)
			encryptor = cipher.encryptor()
			encrypted_data = encryptor.update(pad_data(data)) + encryptor.finalize()
			return encrypted_data
			
		time.sleep(5)
		input_hex = Response
		modified_hex = remove_bytes(input_hex)
		time.sleep(2)
		key_input = "fcbe88cb983a0bcc7919678e34ad5275"
		data_input = modified_hex
		key = bytes.fromhex(key_input)
		data = bytes.fromhex(data_input)
		# Check if the length of key and data is correct
		if len(key) != 16 or len(data) != 16:
			print("Key and data must be 128 bits (16 bytes) long.")
			sys.exit(1)
		# Perform AES encryption
		ciphertext = encrypt(key, data)
		final_value = ciphertext[:16]

		modified_ciphertext = b'\x27\x32' + final_value
		encrypted_request = modified_ciphertext.hex()
		return encrypted_request
			
			#print(modified_ciphertext.hex())
			#resp=canoe_inst.send_diag_request('Rear_Left', encrypted_request)
	canoe_inst = CANoe(py_canoe_log_dir=r'C:\.py_canoe')
	canoe_inst.open(canoe_cfg=r'C:\VENGATESH\SW_11_CanoeConfig\HONDA_SRR6p_SW_11_0_CANOE_CONFIG.cfg')
	canoe_inst.get_canoe_version_info()
	canoe_inst.start_measurement()
	time.sleep(2)

	#Check SW version in Canoe
	global SW_Version_canoe
	canoe_inst.check_signal_online('CAN',1,'ADAS_DATA_VERSION_1','SoftwareVersionA')      
	resp1=canoe_inst.get_signal_value('CAN',1,'ADAS_DATA_VERSION_1','SoftwareVersionA')
	resp2=canoe_inst.get_signal_value('CAN',1,'ADAS_DATA_VERSION_1','SoftwareVersionB')
	resp3=canoe_inst.get_signal_value('CAN',1,'ADAS_DATA_VERSION_1','SoftwareVersionC')
	SW_Version_canoe = f"{int(resp1)}_{format(int(resp2),'x').upper()}_{format(int(resp3),'x').upper()}"
	# print(SW_Version_canoe)
	write_to_excel(3,2,SW_Version_canoe,0)

	#Check Extended session
	respe=canoe_inst.send_diag_request('Rear_Left', '10 03')
	write_to_excel(5,2,respe,0)

	#Check Type 1 security
	canoe_inst.send_diag_request('Rear_Left', '27 01')
	respe=canoe_inst.send_diag_request('Rear_Left', '27 02 02 03 02 03 02 03 02 03 02 03 02 03 02 03 02 03 02 03 02 03 02 03 02 03 02 03 02 03 02 03 02 03')
	write_to_excel(6,2,respe,0)

	#Check Type 5 security
	respV=canoe_inst.send_diag_request('Rear_Left', '27 07')
	response_key=remove_bytes(respV)
	request = Type_V_Security(response_key)
	respe = canoe_inst.send_diag_request('Rear_Left', request)
	write_to_excel(7,2,respe,0)

	#Check Type X security
	resp=canoe_inst.send_diag_request('Rear_Left', '27 31')
	encrypted_request = Type_X_Security(resp)
	respe=canoe_inst.send_diag_request('Rear_Left', encrypted_request)
	write_to_excel(8,2,respe,0)
	
	resp=canoe_inst.send_diag_request('Rear_Left', '104F')
	write_to_excel(16,2,resp,0)
	x=1
	while True: #-----> To send list of DID from excel and read the value and write the response to excel
		
		read_value = read_from_excel(x,1,1)
		if(read_value == None):
			break
		resp=canoe_inst.send_diag_request('Rear_Left', str(read_value))
		write_to_excel(x,2,resp,1)
		x=x+1	
	
	canoe_inst.stop_measurement()
	canoe_state=1
	print("Stopping measurement")
	t32api.T32_Cmd(b"QUIT")
	canoe_inst.quit()   
	return canoe_state    



#***************************************** Main ************************************************************

applications_to_close = ["notepad.exe","EXCEL.EXE","CANoe64.exe"]
close_applications(applications_to_close)

Clear_excel_Report(0)
Clear_excel_Report(1)

Make_file_path = r"C:\VENGATESH\10034766_SW_Honda_SRR6P\software\build\appl.mak"  # Change this to the path of your file

Ini_file_path = r"C:\Flashing\TEST\last_flash_session.ini"


replace_Ini_changes_delta_build = {
    4:"PRE_ERASE OFF",
	}


replace_Ini_changes_full_build = {
    4:"PRE_ERASE ON",
	}
	
#deleting the old out folder first because it will contian the old binary files
while True:
	
	global User_input
	User_input = input("Do you want full build or delta build ! Type F or D     ")
	if((User_input == 'F') or (User_input =='f')):
		
		print("Full build started")
		flash_delay=90
		Comment_Make_File(Make_file_path,"# Include the targets to build",1)
		replace_lines(Ini_file_path,replace_Ini_changes_full_build)
		folder_to_delete = "C:/VENGATESH/10034766_SW_Honda_SRR6P/out"  # Replace with the actual path
		delete_folder(folder_to_delete)
		arguments = ["autosarc", "autosar", "clean", "srr6p_all"]
		break
		
	elif((User_input == 'D') or (User_input =='d')):
		print("Delta build started")
		flash_delay=90
		Comment_Make_File(Make_file_path,"# Include the targets to build",0)
		replace_lines(Ini_file_path,replace_Ini_changes_delta_build)
		arguments = ["srr6p"]
		break
		
	else:
		print("Retry Again")
		sys.exit(1)

#building the ptp
file_path = "C:/VENGATESH/10034766_SW_Honda_SRR6P/software/build"
batch_file_name = "build.bat"

if((User_input == 'F') or (User_input =='f')):

	execute_batch_file(file_path, batch_file_name, arguments)
	
elif((User_input == 'D') or (User_input =='d')):

	log_file = "C:/Flashing/TEST/buildlog.txt"
	file_path = "C:/VENGATESH/10034766_SW_Honda_SRR6P/software/build"
	log_command_output(file_path,arguments, log_file)

# Specify the filename and string to search for
filename1 = "C:/VENGATESH/10034766_SW_Honda_SRR6P/out/build_autosarc_autosar_clean_srr6p_all.log"  # Replace with the actual filename
filename2 = "C:/VENGATESH/10034766_SW_Honda_SRR6P/out/build_autosarc_autosar_clean_srr6p_all.log"  # Replace with the actual filename
filename3 = "C:/VENGATESH/10034766_SW_Honda_SRR6P/out/build_autosar.log"
filename4 = "C:/VENGATESH/10034766_SW_Honda_SRR6P/out/build_srr6p_all.log"
filename5 = "C:/Flashing/TEST/buildlog.txt"

string_to_find1 = "autosar.a built"
string_to_find2 = "srr6p_smc_extflash.ptp built"
string_to_find3 = "srr6p_csum_filled.ptp built"

# Check if the string is present in the file
if ((find_string_in_file(filename1, string_to_find1)) or (find_string_in_file(filename3, string_to_find1))):
    autosar_build_state=1
    print("Autosar build completed\n")  #Autosar build completed sucessfully
else:
    autosar_build_state=0
    print("Autosar build failed\n")
	
if ((find_string_in_file(filename2, string_to_find2)) or (find_string_in_file(filename4, string_to_find2))):
    srr6p_build_state=1
    print("SRR6P build completed\n")   # main build completed sucessfully
	
elif(find_string_in_file(filename5,string_to_find3)):
	srr6p_build_state=1
	print("SRR6P delta build completed\n")   # if the build fails due to any reason then activate a log file to capture the build status 	
else:
    srr6p_build_state=0
    print("SRR6P build failed\n")

Generated_SFT_Folder_delete = "C:/VENGATESH/10034766_SW_Honda_SRR6P/out/srr6p/SensorFlashTool"
delete_folder(Generated_SFT_Folder_delete)

Ram_Rom_usage()


if (autosar_build_state == 1 and srr6p_build_state == 1):

	print("\n**************************** Flashing started ****************************\n")
	Start_flashing(flash_delay)
	Start_Canoe()
    
	
if((flash_state == 1 ) and (canoe_state == 1)):   #-----> Indicates flashing is success and creates SW release package

	# First delete the files inside the base path

	folder_to_empty = "C:/VENGATESH/Base"  # Replace with the actual folder path
	delete_folder_contents(folder_to_empty)

	time.sleep(2)
	#create folder to copy files for lauterbach
	SW_Version = f"SW_{SW_Version_canoe}"

	Base_path = "C:/VENGATESH/Base"  # Replace with the actual path
	folder_name_lauterbach = SW_Version  # Replace with the desired folder name
	create_folder(Base_path, folder_name_lauterbach)


	#copy  files for lauterbach

	source_folders1 = ["C:/VENGATESH/10034766_SW_Honda_SRR6P/out/srr6p","C:/VENGATESH/10034766_SW_Honda_SRR6P/out/pbl","C:/VENGATESH/10034766_SW_Honda_SRR6P/out/sbl"]
	destination_folder1 = f"C:/VENGATESH/Base/{folder_name_lauterbach}"
	filename_list1 = ["srr6p.cmd", "srr6p.elf", "srr6p.map", "srr6p.s19", "srr6p_csum.ptp", "srr6p_csum_extflash.ptp", "srr6p_csum_filled.ptp", "srr6p_memory_stats.txt","srr6p_smc_extflash.ptp","srr6p_usc_extflash.ptp","pbl.cmd", "pbl.elf","pbl.map","pbl_csum_filled.ptp", "sbl.cmd", "sbl.elf", "sbl.map","sbl_csum_filled.ptp"]
	copy_files(source_folders1, destination_folder1, filename_list1)



	#create folder to copy files for SFT

	Base_path = "C:/VENGATESH/Base"  # Replace with the actual path
	folder_name_SFT = f"SensorFlashTool_Honda_{SW_Version}"  # Replace with the desired folder name
	create_folder(Base_path, folder_name_SFT)

	#copy  files for SFT
	source_folders2 = ["C:/VENGATESH/10034766_SW_Honda_SRR6P/out/srr6p", "C:/VENGATESH/10034766_SW_Honda_SRR6P/out/pbl", "C:/VENGATESH/10034766_SW_Honda_SRR6P/out/sbl","C:/VENGATESH/Release_Folder/SensorFlashTool_Honda_base", ]
	destination_folder2 = f"C:/VENGATESH/Base/{folder_name_SFT}"
	filename_list2 = ["pbl_csum_filled.ptp","sbl_csum_filled.ptp","SensorFlashTool.exe","SensorFlashTool_GEN5_APP_SRR6P.xml","SensorFlashTool_GEN5_APP_SRR6P_VTV_Only.xml","SensorFlashTool_GEN5_BL_SRR6P.xml","SensorFlashTool_GEN5_SPC_SRR6P.xml","SensorFlashTool_GEN5_USC_SRR6P.xml","USC_3_22_0_0.ptp","USC_3_22_0_1_009.ptp","vxlapi.dll","SensorFlashTool.pdf","srr6p_csum_extflash.ptp","srr6p_smc_extflash.ptp"]
	copy_files(source_folders2, destination_folder2, filename_list2)
	
	# This function will ZIP the folders based on the path given
	root_dir = "C:/VENGATESH/Base"  # Replace with the actual path
	zip_folders(root_dir)


	# After Zipping the contents move the contents to release folder
	global Final_SW_folder_name
	Final_SW_folder_name= f"SW_{SW_Version_canoe}"
	source_path = "C:/VENGATESH/Base"  # Replace with the actual source path
	destination_path = f"C:/VENGATESH/Release_Folder/{Final_SW_folder_name}"  # Replace with the actual destination path
	copy_folders_and_files(source_path, destination_path)
	
	#Fill the RCR checklist
	Fill_Checklist(Final_SW_folder_name)
	Copy_Excel()
	Rename_Excel()
	
	sys.exit(0)
	
else:
	
	sys.exit(1)